import openpyxl, json
from openpyxl.utils import get_column_letter
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                     HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                     FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

row_map = json.load(open("/home/claude/cntb/row_map.json"))
A = row_map["ASSUMP_START"]
POS_ASTHMA_PH2, POS_ASTHMA_CONV, POS_COPD_APPR, PEAK_PTS_ASTHMA, PEAK_PTS_COPD, NET_PRICE, PEAK_PENETRATION, \
LAUNCH_ASTHMA, LAUNCH_COPD, YEARS_TO_PEAK, LOE_YEAR, COGS_PCT, SGA_PCT, WACC, CHINA_MS_PCT, CHINA_ROY_NPV = \
[A+i for i in range(16)]
CA = "'Cover & Assumptions'"
def af(row): return f"{CA}!$F${row}"

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb["rNPV Model"]

N = row_map["N_YEARS"]; FIRST_COL_IDX = row_map["FIRST_COL_IDX"]
def yc(i): return get_column_letter(FIRST_COL_IDX + i)
YEAR_ROW = row_map["YEAR_ROW"]

r = row_map["NEXT_ROW_AFTER_ASTHMA"]

section_bar(ws, r, "PROGRAM 2 — ACUTE EXACERBATIONS OF COPD (US)"); r += 2

set_cell(ws, f"B{r}", "Years since launch", BLACK)
YSL_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}${YEAR_ROW}-{af(LAUNCH_COPD)}", BLACK, num_fmt=NUM, align="center")
r += 1

set_cell(ws, f"B{r}", "Penetration of eligible population", BLACK)
PEN_C_ROW = r
for i in range(N):
    col = yc(i)
    f = (f"=IF({col}{YSL_C_ROW}<0,0,IF({col}{YSL_C_ROW}>={af(LOE_YEAR)}-{af(LAUNCH_COPD)},"
         f"{af(PEAK_PENETRATION)}*0.15*0.8,MIN({af(PEAK_PENETRATION)}*0.8,{af(PEAK_PENETRATION)}*0.8*{col}{YSL_C_ROW}/{af(YEARS_TO_PEAK)})))")
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=PCT, align="center")
r += 1
set_cell(ws, "X"+str(PEN_C_ROW), "COPD peak penetration set at 80% of the Asthma assumption, reflecting a more crowded/skeptical prescriber base for a first-in-class acute-exacerbation biologic in COPD specifically.", NOTE, border=False, wrap=True)

set_cell(ws, f"B{r}", "Treated patients (000s)", BLACK)
PTS_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={af(PEAK_PTS_COPD)}*{col}{PEN_C_ROW}", BLACK, num_fmt=NUM1, align="center")
r += 1

set_cell(ws, f"B{r}", "Unrisked US net revenue ($mm)", BLACK)
REV_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{PTS_C_ROW}*1000*{af(NET_PRICE)}/1000000", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "COGS ($mm)", BLACK)
COGS_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=-{col}{REV_C_ROW}*{af(COGS_PCT)}", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "SG&A ($mm, incremental — shared field force w/ Asthma)", BLACK)
SGA_C_ROW = r
for i in range(N):
    col = yc(i)
    f = (f"=IF(AND({col}{YSL_C_ROW}<0,{col}{YSL_C_ROW}>=-2),-4,"
         f"IF({col}{REV_C_ROW}>0,-{col}{REV_C_ROW}*{af(SGA_PCT)}*0.7,0))")
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=CUR, align="center")
r += 1
set_cell(ws, "X"+str(SGA_C_ROW), "Incremental SG&A only, at 70% of the standalone rate — COPD launches into an already-built acute-care commercial infrastructure from the Asthma launch.", NOTE, border=False, wrap=True)

set_cell(ws, f"B{r}", "Lifecycle R&D ($mm, ~5% of revenue)", BLACK)
RD_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=-{col}{REV_C_ROW}*0.05", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Pre-launch Phase 3 program cost ($mm)", BLACK)
PH3_C_ROW = r
for i in range(N):
    col = yc(i)
    f = f"=IF(AND({col}{YSL_C_ROW}<0,{col}{YSL_C_ROW}>=-3),-35,0)"
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Unrisked pre-tax cash flow ($mm)", BOLD_BLACK)
PRETAX_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{REV_C_ROW}+{col}{COGS_C_ROW}+{col}{SGA_C_ROW}+{col}{RD_C_ROW}+{col}{PH3_C_ROW}", BOLD_BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Effective cash tax rate", BLACK)
TAX_C_ROW = r
YEARS = list(range(2026, 2026+N))
for i in range(N):
    col = yc(i)
    y = YEARS[i]
    set_cell(ws, f"{col}{r}", 0.0 if y < 2035 else 0.21, BLUE, FILL_INPUT, PCT, align="center")
r += 1

set_cell(ws, f"B{r}", "Unrisked after-tax cash flow ($mm)", BOLD_BLACK)
AT_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=IF({col}{PRETAX_C_ROW}>0,{col}{PRETAX_C_ROW}*(1-{col}{TAX_C_ROW}),{col}{PRETAX_C_ROW})", BOLD_BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Cumulative PoS to approval — COPD", BLACK)
COMBPOS_C_ROW = r
set_cell(ws, f"C{r}", f"={af(POS_COPD_APPR)}", BLACK, num_fmt=PCT, align="center")
ws.merge_cells(f"C{r}:{yc(N-1)}{r}")
r += 1

set_cell(ws, f"B{r}", "Risk-adjusted after-tax cash flow ($mm)", BOLD_BLACK, FILL_LIGHT)
RISK_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{AT_C_ROW}*$C${COMBPOS_C_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Discount factor", BLACK)
DISC_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=1/(1+{af(WACC)})^({col}${YEAR_ROW}-2026+0.5)", BLACK, num_fmt='0.000', align="center")
r += 1

set_cell(ws, f"B{r}", "PV of risk-adjusted cash flow ($mm)", BLACK)
PV_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{RISK_C_ROW}*{col}{DISC_C_ROW}", BLACK, num_fmt=CUR, align="center")
r += 2

set_cell(ws, f"B{r}", "rNPV — COPD program ($mm)", BOLD_BLACK, FILL_LIGHT)
RNPV_COPD_TOTAL_ROW = r
set_cell(ws, f"C{r}", f"=SUM({yc(0)}{PV_C_ROW}:{yc(N-1)}{PV_C_ROW})", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
r += 2

# ---------------- SUMMARY ----------------
section_bar(ws, r, "rNPV SUMMARY & PER-SHARE VALUE"); r += 2

RNPV_ASTHMA_TOTAL_ROW = row_map["RNPV_ASTHMA_TOTAL_ROW"]

summary_rows = []
set_cell(ws, f"B{r}", "rNPV — Asthma program ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=C{RNPV_ASTHMA_TOTAL_ROW}", BLACK, num_fmt=CUR)
ASTHMA_SUMM_ROW = r; r += 1

set_cell(ws, f"B{r}", "rNPV — COPD program ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=C{RNPV_COPD_TOTAL_ROW}", BLACK, num_fmt=CUR)
COPD_SUMM_ROW = r; r += 1

set_cell(ws, f"B{r}", "China milestones, risk/time-weighted ($mm)", BLACK)
set_cell(ws, f"C{r}", f"={af(CHINA_MS_PCT)}*110", BLACK, num_fmt=CUR)
CHINA_MS_SUMM_ROW = r; r += 1

set_cell(ws, f"B{r}", "China royalty stream, risk-adjusted NPV ($mm)", BLACK)
set_cell(ws, f"C{r}", f"={af(CHINA_ROY_NPV)}", BLACK, num_fmt=CUR)
CHINA_ROY_SUMM_ROW = r; r += 1

set_cell(ws, f"B{r}", "Sum of program rNPVs ($mm)", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=SUM(C{ASTHMA_SUMM_ROW}:C{CHINA_ROY_SUMM_ROW})", BOLD_BLACK, num_fmt=CUR)
SUM_RNPV_ROW = r; r += 1

set_cell(ws, f"B{r}", "Plus: net cash ($mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['CASH_ROW']}-'Cover & Assumptions'!C{row_map['DEBT_ROW']}", GREEN, num_fmt=CUR)
NETCASH_ROW = r; r += 1

set_cell(ws, f"B{r}", "Less: 24 months of further operating burn to/through data readouts ($mm)", BLACK)
set_cell(ws, f"C{r}", "='Balance Sheet & Cash Runway'!C35*8", GREEN, num_fmt=CUR)
set_cell(ws, f"G{r}", "Conservatism check: rNPV is a going-concern value, but near-term financing dilution is real and not otherwise captured above (no future equity raise is modeled). We net out ~8 quarters of burn as a simple placeholder for the dilution/cash drag between now and Ph3 funding — remove this line if you'd rather see the undiluted rNPV and treat financing risk qualitatively.", NOTE, border=False, wrap=True)
BURN_DRAG_ROW = r; r += 1

set_cell(ws, f"B{r}", "Total equity value ($mm)", BOLD_BLACK, FILL_LIGHT)
set_cell(ws, f"C{r}", f"=C{SUM_RNPV_ROW}+C{NETCASH_ROW}-C{BURN_DRAG_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
TOTAL_EQ_ROW = r; r += 1

set_cell(ws, f"B{r}", "Fully-diluted shares (mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['FD_ROW']}", GREEN, num_fmt=NUM1)
FD_ROW_LOCAL = r; r += 2

set_cell(ws, f"B{r}", "rNPV per share ($)", BOLD_BLACK, FILL_YELLOW)
set_cell(ws, f"C{r}", f"=C{TOTAL_EQ_ROW}/C{FD_ROW_LOCAL}", BOLD_BLACK, FILL_YELLOW, num_fmt=CUR2)
RNPV_PER_SHARE_ROW = r; r += 2

set_cell(ws, f"B{r}", ("Reality check: this rNPV blends a still-binary Phase 2 readout (Sept 2026) with a full 20-year commercial forecast built on several UNKNOWN inputs (pricing, peak share, patent life). "
                       "Treat the Low/Base/High spread, not the Base point estimate, as the honest output of this tab."), BLACK, border=False, wrap=True)
ws.merge_cells(f"B{r}:{yc(8)}{r+2}")
ws.row_dimensions[r].height = 55

ws.freeze_panes = "C6"
wb.save("/home/claude/cntb/model.xlsx")

row_map.update(dict(
    YSL_C_ROW=YSL_C_ROW, PEN_C_ROW=PEN_C_ROW, PTS_C_ROW=PTS_C_ROW, REV_C_ROW=REV_C_ROW, COGS_C_ROW=COGS_C_ROW,
    SGA_C_ROW=SGA_C_ROW, RD_C_ROW=RD_C_ROW, PH3_C_ROW=PH3_C_ROW, PRETAX_C_ROW=PRETAX_C_ROW, TAX_C_ROW=TAX_C_ROW,
    AT_C_ROW=AT_C_ROW, COMBPOS_C_ROW=COMBPOS_C_ROW, RISK_C_ROW=RISK_C_ROW, DISC_C_ROW=DISC_C_ROW, PV_C_ROW=PV_C_ROW,
    RNPV_COPD_TOTAL_ROW=RNPV_COPD_TOTAL_ROW, RNPV_PER_SHARE_ROW=RNPV_PER_SHARE_ROW, TOTAL_EQ_ROW=TOTAL_EQ_ROW,
    SUM_RNPV_ROW=SUM_RNPV_ROW,
))
json.dump(row_map, open("/home/claude/cntb/row_map.json","w"))
print("rNPV per-share row:", RNPV_PER_SHARE_ROW)
