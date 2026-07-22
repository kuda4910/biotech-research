import openpyxl, json
row_map = json.load(open("/home/claude/cntb/row_map.json"))
wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb["Cover & Assumptions"]
A = row_map["ASSUMP_START"]
POS_COPD_APPR = A+2   # 44
NET_PRICE = A+5       # 47
PEAK_PENETRATION = A+6  # 48

# COPD PoS -> lower than Asthma combined PoS (consistent w/ narrative that COPD is riskier/slower)
ws[f"C{POS_COPD_APPR}"] = 0.18
ws[f"D{POS_COPD_APPR}"] = 0.28
ws[f"E{POS_COPD_APPR}"] = 0.40

# Net price per patient - trim down, still an estimate
ws[f"C{NET_PRICE}"] = 6000
ws[f"D{NET_PRICE}"] = 9000
ws[f"E{NET_PRICE}"] = 13000

# Peak penetration - trim down slightly
ws[f"C{PEAK_PENETRATION}"] = 0.12
ws[f"D{PEAK_PENETRATION}"] = 0.20
ws[f"E{PEAK_PENETRATION}"] = 0.30

wb.save("/home/claude/cntb/model.xlsx")
print("patched assumptions")
