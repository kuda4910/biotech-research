import openpyxl
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                     HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                     FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb["Cover & Assumptions"]

r = 19
section_bar(ws, r, "PROGRAM STATUS & PIPELINE"); r += 2

rows_prog = [
 ("Program 1 — US/global:", "Ph2 SEABREEZE STAT Asthma (n=160, acute exacerbations, EOS≥300); enrollment complete 17-Jun-2026; primary endpoint = 28-day treatment-failure rate; key secondary = post-BD FEV1 at Wk1.", "Topline expected early Sept 2026 (company guidance, 17-Jun-2026 PR)."),
 ("Program 2 — US/global:", "Ph2 SEABREEZE STAT COPD (n=160, acute exacerbations, EOS≥300); enrollment substantially complete ~Jun 2026.", "Topline expected shortly after Asthma readout — company has not given a firm date; we ESTIMATE Q4 2026."),
 ("Supportive data:", "Prior global Ph2 chronic moderate-to-severe asthma trial (positive, published AJRCCM) + Ph1 IV rademikibart (rapid FEV1 improvement within 24h, sustained through ~4 wks).", "Completed; de-risks mechanism, not the acute-exacerbation endpoint specifically."),
 ("China (Simcere license):", "Simcere holds exclusive rights to rademikibart in Greater China, all indications. Ph3 AD trial in China met all primary/secondary endpoints; NDA for AD submitted to China's regulator (2025/26).", "CNTB eligible for ~$110mm remaining dev/reg/commercial milestones + tiered low-double-digit royalties on Greater China net sales."),
 ("DMC interim review:", "Independent DMC completed pre-specified interim efficacy review of both SEABREEZE STAT studies (23-Apr-2026): no safety concerns, no change to sample size recommended.", "Positive procedural signal; not a read on efficacy magnitude."),
]
for label, detail, note in rows_prog:
    set_cell(ws, f"B{r}", label, BOLD_BLACK, border=False)
    set_cell(ws, f"C{r}", detail, BLACK, border=False, wrap=True)
    ws.merge_cells(f"C{r}:E{r}")
    ws.row_dimensions[r].height = 42
    set_cell(ws, f"G{r}", note, NOTE, border=False, wrap=True)
    r += 1

r += 1
section_bar(ws, r, "KEY UPCOMING CATALYSTS / DATES"); r += 2
dates = [
    ("Q2 2026 earnings", "12-Aug-2026", "Confirmed (StockAnalysis.com earnings calendar)."),
    ("SEABREEZE STAT Asthma topline", "Early Sept 2026", "Company guidance, 17-Jun-2026 press release."),
    ("SEABREEZE STAT COPD topline", "Q4 2026 (EST.)", "ESTIMATE — company says 'shortly after' asthma; no firm date given."),
    ("FDA Phase 3 alignment meeting", "H2 2026 / H1 2027 (EST.)", "ESTIMATE — company states intent to 'move quickly' post-data; no confirmed date."),
    ("Cash runway (per management)", "Into 2H 2027", "Per Q1'26 10-Q / 8-K, based on current operating plan; excludes any new financing."),
]
for label, date, note in dates:
    set_cell(ws, f"B{r}", label, BOLD_BLACK)
    set_cell(ws, f"C{r}", date, BLACK)
    ws.merge_cells(f"C{r}:D{r}")
    set_cell(ws, f"G{r}", note, NOTE, border=False, wrap=True)
    r += 1

r += 1
SCEN_ROW = r
section_bar(ws, r, "SCENARIO SELECTOR"); r += 2
set_cell(ws, f"B{r}", "Active scenario (drives every downstream tab):", BOLD_BLACK)
SCEN_CELL_ROW = r
set_cell(ws, f"C{r}", "Base", BLUE, FILL_YELLOW, align="center")
dv = DataValidation(type="list", formula1='"Low,Base,High"', allow_blank=False)
ws.add_data_validation(dv)
dv.add(ws[f"C{r}"])
set_cell(ws, f"G{r}", "Change this cell to Low / Base / High — every assumption below, and the rNPV/DCF/relative-valuation tabs, re-flow automatically.", NOTE, border=False, wrap=True)
r += 2

SCEN_REF = f"$C${SCEN_CELL_ROW}"

section_bar(ws, r, "MASTER ASSUMPTIONS (Low / Base / High, with Selected value flowing downstream)"); r += 2
hdr_row = r
headers = ["Assumption", "Low", "Base", "High", "Selected", "Source / Basis"]
for i, h in enumerate(headers):
    col = chr(ord('B') + i)
    set_cell(ws, f"{col}{hdr_row}", h, HEADER, FILL_HEADER, align="center")
r += 1

ASSUMP_START = r
assumptions = [
 ("PoS: SEABREEZE Asthma Ph2 — positive topline", 0.40, 0.60, 0.75, PCT,
  "ESTIMATE/JUDGMENT. Mechanism de-risked by validated IL-4Rα target (Dupixent), positive Ph1 IV & prior chronic-asthma Ph2 (AJRCCM), and a clean DMC interim review — but the acute-exacerbation endpoint (28-day treatment failure) has not been tested for this mechanism before. Treat as a genuine binary-risk estimate, not a fact."),
 ("PoS: Ph2(+)→US Approval, Asthma (execution+FDA risk)", 0.35, 0.50, 0.65, PCT,
  "ESTIMATE. Reflects typical Ph2→approval conversion for validated-MOA respiratory biologics (industry base rates for Ph2 immunology assets run ~30-55%)."),
 ("PoS: US Approval, COPD (own binary + asthma read-through)", 0.25, 0.40, 0.55, PCT,
  "ESTIMATE — COPD approval history for Th2/IL-4 biologics is mixed and slower (e.g., Dupixent's COPD approval came years after asthma, in a narrower population). We discount COPD below asthma for this reason."),
 ("US peak eligible patients — Asthma (000s/yr)", 350, 450, 550, NUM,
  "Bottom-up: company discloses >1mm US ED visits/yr for acute asthma exacerbation; we apply ~35-55% for EOS≥300/FeNO≥25 (biomarker-eligible) and net for repeat-visit patients. GENUINE ESTIMATE — no third-party epi study cited by company."),
 ("US peak eligible patients — COPD (000s/yr)", 400, 550, 700, NUM,
  "Bottom-up: CDC-cited ~1.5mm annual US COPD-exacerbation hospitalizations/ED visits; ESTIMATE ~35-50% EOS≥300 eligible, netted for repeat visits."),
 ("Net price per treated patient per year ($)", 8000, 12000, 16000, CUR,
  "UNKNOWN / GENUINE ESTIMATE — company has given no pricing guidance. Proxy: episodic/acute specialty-biologic pricing analogues, net of gross-to-net. This is the single largest swing factor in the model."),
 ("Peak penetration of eligible population", 0.15, 0.25, 0.35, PCT,
  "ESTIMATE. No approved biologic exists in the acute-exacerbation setting today (differentiated), but point-of-care/ED biologic administration is a novel commercial model with real adoption-curve risk."),
 ("US launch year — Asthma", 2032, 2031, 2030, NUM,
  "ESTIMATE from data timing: Ph2 data Sep-2026 (Base) → ~2yr Ph3 → filing/review. High = faster path (e.g., breakthrough designation); Low = slower."),
 ("US launch year — COPD", 2033, 2032, 2031, NUM,
  "ESTIMATE — assumes ~1yr behind Asthma given later Ph2 data and typically slower COPD regulatory path."),
 ("Years from launch to peak sales", 6, 5, 4, NUM,
  "ESTIMATE — typical biologic launch curve in a new site-of-care (ED/urgent care) commercial model."),
 ("Loss of exclusivity year (biosimilar entry)", 2038, 2040, 2042, NUM,
  "ESTIMATE — based on typical biologics exclusivity (12yr BLA) plus composition-of-matter patent term; we do not have CNTB's actual patent expiry schedule from a filing — VERIFY against the 10-K IP section."),
 ("COGS, % of net sales", 0.18, 0.15, 0.12, PCT,
  "ESTIMATE — typical mAb manufacturing cost of goods at commercial scale."),
 ("Peak SG&A, % of net sales", 0.35, 0.28, 0.22, PCT,
  "ESTIMATE — reflects a niche, high-touch ED/hospital commercial model rather than a broad primary-care biologic launch."),
 ("Discount rate / WACC", 0.14, 0.125, 0.11, PCT,
  "ESTIMATE — standard range for a single-asset, pre-revenue (ex-China) clinical biotech; Low scenario uses the higher discount rate (more conservative)."),
 ("China milestone realization (% of $110mm, PV-weighted)", 0.20, 0.35, 0.50, PCT,
  "ESTIMATE of probability/timing-weighted realization of the remaining ~$110mm Simcere milestones, given AD NDA is filed in China but commercial/regulatory milestones remain contingent on Simcere's execution, which CNTB does not control."),
 ("China royalty NPV ($mm, risk-adjusted, standalone)", 15, 30, 50, CUR,
  "ESTIMATE — rough risk-adjusted NPV of low-double-digit royalties on Simcere's Greater China rademikibart sales (AD + potential future indications). We do not have Simcere's internal sales forecasts; this is a placeholder order-of-magnitude, not a bottom-up China epi model."),
]

for name, lo, base, hi, fmt, note in assumptions:
    set_cell(ws, f"B{r}", name, BLACK, wrap=True)
    set_cell(ws, f"C{r}", lo, BLUE, FILL_INPUT, fmt, align="center")
    set_cell(ws, f"D{r}", base, BLUE, FILL_INPUT, fmt, align="center")
    set_cell(ws, f"E{r}", hi, BLUE, FILL_INPUT, fmt, align="center")
    set_cell(ws, f"F{r}", f'=IF({SCEN_REF}="Low",C{r},IF({SCEN_REF}="Base",D{r},E{r}))', BOLD_BLACK, fill=FILL_LIGHT, num_fmt=fmt, align="center")
    set_cell(ws, f"G{r}", note, NOTE, border=False, wrap=True)
    ws.row_dimensions[r].height = 44
    r += 1
ASSUMP_END = r - 1

ws.freeze_panes = "B7"

wb.save("/home/claude/cntb/model.xlsx")

# Print row map for reference by later scripts
row_map = {
    "PRICE_ROW": 10, "SHARES_ROW": 11, "FD_ROW": 12, "MKTCAP_ROW": 13, "CASH_ROW": 15, "DEBT_ROW": 16, "EV_ROW": 17,
    "SCEN_CELL_ROW": SCEN_CELL_ROW, "ASSUMP_START": ASSUMP_START, "ASSUMP_END": ASSUMP_END,
}
print(row_map)
import json
with open("/home/claude/cntb/row_map.json","w") as f:
    json.dump(row_map, f)
