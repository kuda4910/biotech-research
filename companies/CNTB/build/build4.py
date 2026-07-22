import openpyxl, json
from openpyxl.utils import get_column_letter
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                     HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                     FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

row_map = json.load(open("/home/claude/cntb/row_map.json"))
A = row_map["ASSUMP_START"]  # 42
# assumption row offsets
POS_ASTHMA_PH2   = A+0   # 42
POS_ASTHMA_CONV  = A+1   # 43
POS_COPD_APPR    = A+2   # 44
PEAK_PTS_ASTHMA  = A+3   # 45
PEAK_PTS_COPD    = A+4   # 46
NET_PRICE        = A+5   # 47
PEAK_PENETRATION = A+6   # 48
LAUNCH_ASTHMA    = A+7   # 49
LAUNCH_COPD      = A+8   # 50
YEARS_TO_PEAK    = A+9   # 51
LOE_YEAR         = A+10  # 52
COGS_PCT         = A+11  # 53
SGA_PCT          = A+12  # 54
WACC             = A+13  # 55
CHINA_MS_PCT     = A+14  # 56
CHINA_ROY_NPV    = A+15  # 57

CA = "'Cover & Assumptions'"
def af(row):
    return f"{CA}!$F${row}"

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb.create_sheet("rNPV Model")
ws.sheet_view.showGridLines = False

YEARS = list(range(2026, 2046))  # 20 years, 2026..2045
N = len(YEARS)
FIRST_COL_IDX = 3  # column C
def yc(i):
    """column letter for year index i (0-based)"""
    return get_column_letter(FIRST_COL_IDX + i)

widths = {"A": 3, "B": 46}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for i in range(N):
    ws.column_dimensions[yc(i)].width = 11

set_cell(ws, "B2", "rNPV Model — Rademikibart, Acute Exacerbations of Asthma & COPD", TITLE, border=False)
set_cell(ws, "B3", "Bottom-up, epidemiology-driven, probability-of-success risk-adjusted. This should be the most conservative, defensible number in the file — every unrisked dollar here is explicitly discounted by the odds of getting there.", SUBTITLE, border=False, wrap=True)
ws.merge_cells(f"B3:{yc(N-1)}3")

r = 5
set_cell(ws, f"B{r}", "Fiscal Year", BOLD_BLACK, FILL_HEADER, align="center")
ws[f"B{r}"].font = HEADER
for i, y in enumerate(YEARS):
    set_cell(ws, f"{yc(i)}{r}", y, HEADER, FILL_HEADER, align="center")
YEAR_ROW = r
r += 2

# ---------------- ASTHMA BLOCK ----------------
ws.column_dimensions["X"].width = 55
section_bar(ws, r, "PROGRAM 1 — ACUTE EXACERBATIONS OF ASTHMA (US)"); r += 2

set_cell(ws, f"B{r}", "Years since launch", BLACK)
YSL_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}${YEAR_ROW}-{af(LAUNCH_ASTHMA)}", BLACK, num_fmt=NUM, align="center")
r += 1

set_cell(ws, f"B{r}", "Penetration of eligible population", BLACK)
PEN_A_ROW = r
for i in range(N):
    col = yc(i)
    f = (f"=IF({col}{YSL_A_ROW}<0,0,IF({col}{YSL_A_ROW}>={af(LOE_YEAR)}-{af(LAUNCH_ASTHMA)},"
         f"{af(PEAK_PENETRATION)}*0.15,MIN({af(PEAK_PENETRATION)},{af(PEAK_PENETRATION)}*{col}{YSL_A_ROW}/{af(YEARS_TO_PEAK)})))")
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=PCT, align="center")
r += 1

set_cell(ws, f"B{r}", "Treated patients (000s)", BLACK)
PTS_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={af(PEAK_PTS_ASTHMA)}*{col}{PEN_A_ROW}", BLACK, num_fmt=NUM1, align="center")
r += 1

set_cell(ws, f"B{r}", "Unrisked US net revenue ($mm)", BLACK)
REV_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{PTS_A_ROW}*1000*{af(NET_PRICE)}/1000000", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "COGS ($mm)", BLACK)
COGS_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=-{col}{REV_A_ROW}*{af(COGS_PCT)}", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "SG&A ($mm)", BLACK)
SGA_A_ROW = r
for i in range(N):
    col = yc(i)
    # pre-launch buildout: 2 yrs before launch, $8mm/yr commercial infra spend
    f = (f"=IF(AND({col}{YSL_A_ROW}<0,{col}{YSL_A_ROW}>=-2),-8,"
         f"IF({col}{REV_A_ROW}>0,-{col}{REV_A_ROW}*{af(SGA_PCT)},0))")
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Lifecycle R&D ($mm, ~5% of revenue)", BLACK)
RD_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=-{col}{REV_A_ROW}*0.05", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Pre-launch Phase 3 program cost ($mm)", BLACK)
PH3_A_ROW = r
for i in range(N):
    col = yc(i)
    # Ph3 spend concentrated in years -3 to -1 relative to launch, ~$40mm/yr unrisked
    f = f"=IF(AND({col}{YSL_A_ROW}<0,{col}{YSL_A_ROW}>=-3),-40,0)"
    set_cell(ws, f"{col}{r}", f, BLACK, num_fmt=CUR, align="center")
set_cell(ws, "B"+str(PH3_A_ROW+0), "Pre-launch Phase 3 program cost ($mm)", BLACK)
r += 1

set_cell(ws, f"B{r}", "Unrisked pre-tax cash flow ($mm)", BOLD_BLACK)
PRETAX_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=SUM({col}{REV_A_ROW}:{col}{PH3_A_ROW})-{col}{REV_A_ROW}", BOLD_BLACK, num_fmt=CUR, align="center")
    # NOTE: sum includes REV_A_ROW twice fix below
r += 1
# Fix: recompute properly as REV+COGS+SGA+RD+PH3 (not double counting)
for i in range(N):
    col = yc(i)
    ws[f"{col}{PRETAX_A_ROW}"] = f"={col}{REV_A_ROW}+{col}{COGS_A_ROW}+{col}{SGA_A_ROW}+{col}{RD_A_ROW}+{col}{PH3_A_ROW}"

set_cell(ws, f"B{r}", "Effective cash tax rate", BLACK)
TAX_A_ROW = r
for i in range(N):
    col = yc(i)
    y = YEARS[i]
    set_cell(ws, f"{col}{r}", 0.0 if y < 2034 else 0.21, BLUE, FILL_INPUT, PCT, align="center")
r += 1
set_cell(ws, "X"+str(TAX_A_ROW), "ESTIMATE — assumes NOLs shelter income through 2033, then statutory 21% federal rate. We do not have CNTB's actual NOL balance/carryforward schedule (10-K tax footnote) — verify.", NOTE, border=False, wrap=True)

set_cell(ws, f"B{r}", "Unrisked after-tax cash flow ($mm)", BOLD_BLACK)
AT_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=IF({col}{PRETAX_A_ROW}>0,{col}{PRETAX_A_ROW}*(1-{col}{TAX_A_ROW}),{col}{PRETAX_A_ROW})", BOLD_BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Cumulative PoS to approval — Asthma", BLACK)
COMBPOS_A_ROW = r
set_cell(ws, f"C{r}", f"={af(POS_ASTHMA_PH2)}*{af(POS_ASTHMA_CONV)}", BLACK, num_fmt=PCT, align="center")
ws.merge_cells(f"C{r}:{yc(N-1)}{r}")
r += 1

set_cell(ws, f"B{r}", "Risk-adjusted after-tax cash flow ($mm)", BOLD_BLACK, FILL_LIGHT)
RISK_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{AT_A_ROW}*$C${COMBPOS_A_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Discount factor", BLACK)
DISC_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=1/(1+{af(WACC)})^({col}${YEAR_ROW}-2026+0.5)", BLACK, num_fmt='0.000', align="center")
r += 1

set_cell(ws, f"B{r}", "PV of risk-adjusted cash flow ($mm)", BLACK)
PV_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{RISK_A_ROW}*{col}{DISC_A_ROW}", BLACK, num_fmt=CUR, align="center")
r += 2

set_cell(ws, f"B{r}", "rNPV — Asthma program ($mm)", BOLD_BLACK, FILL_LIGHT)
RNPV_ASTHMA_TOTAL_ROW = r
set_cell(ws, f"C{r}", f"=SUM({yc(0)}{PV_A_ROW}:{yc(N-1)}{PV_A_ROW})", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
r += 2

wb.save("/home/claude/cntb/model.xlsx")
row_map.update(dict(
    YEAR_ROW=YEAR_ROW, YSL_A_ROW=YSL_A_ROW, PEN_A_ROW=PEN_A_ROW, PTS_A_ROW=PTS_A_ROW, REV_A_ROW=REV_A_ROW,
    COGS_A_ROW=COGS_A_ROW, SGA_A_ROW=SGA_A_ROW, RD_A_ROW=RD_A_ROW, PH3_A_ROW=PH3_A_ROW, PRETAX_A_ROW=PRETAX_A_ROW,
    TAX_A_ROW=TAX_A_ROW, AT_A_ROW=AT_A_ROW, COMBPOS_A_ROW=COMBPOS_A_ROW, RISK_A_ROW=RISK_A_ROW, DISC_A_ROW=DISC_A_ROW,
    PV_A_ROW=PV_A_ROW, RNPV_ASTHMA_TOTAL_ROW=RNPV_ASTHMA_TOTAL_ROW, NEXT_ROW_AFTER_ASTHMA=r,
    N_YEARS=N, FIRST_COL_IDX=FIRST_COL_IDX,
))
json.dump(row_map, open("/home/claude/cntb/row_map.json","w"))
print(r)
