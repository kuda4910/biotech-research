import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

FONT = "Arial"

BLUE = Font(name=FONT, color="0000FF", size=10)
BLACK = Font(name=FONT, color="000000", size=10)
GREEN = Font(name=FONT, color="008000", size=10)
BOLD = Font(name=FONT, bold=True, size=10)
BOLD_BLACK = Font(name=FONT, bold=True, color="000000", size=10)
TITLE = Font(name=FONT, bold=True, size=16, color="1F3864")
SUBTITLE = Font(name=FONT, italic=True, size=10, color="595959")
HEADER = Font(name=FONT, bold=True, size=11, color="FFFFFF")
SECTION = Font(name=FONT, bold=True, size=12, color="FFFFFF")
NOTE = Font(name=FONT, italic=True, size=9, color="808080")
RED = Font(name=FONT, color="FF0000", size=10)

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_SECTION = PatternFill("solid", fgColor="2E5395")
FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
FILL_LIGHT = PatternFill("solid", fgColor="D9E1F2")
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%'
MULT = '0.0x'
NUM = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'

def set_cell(ws, coord, value=None, font=None, fill=None, num_fmt=None, align=None, border=True, wrap=False, comment=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if num_fmt:
        c.number_format = num_fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    elif wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    if border:
        c.border = BORDER
    if comment:
        c.comment = Comment(comment, "Analyst")
    return c

def section_bar(ws, row, text, col_start=1, col_end=8):
    for col in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_SECTION
        cell.font = SECTION
    ws.cell(row=row, column=col_start).value = text

def sel_formula(cell_low, cell_high_range_ref, scenario_cell="$B$29"):
    pass

wb = Workbook()
ws = wb.active
ws.title = "Cover & Assumptions"

# Column widths
widths = {"A": 3, "B": 42, "C": 13, "D": 13, "E": 13, "F": 14, "G": 45, "H": 3}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.sheet_view.showGridLines = False

# ---- Title block ----
set_cell(ws, "B2", "Connect Biopharma Holdings Limited (NASDAQ: CNTB)", TITLE, border=False)
ws.merge_cells("B2:G2")
set_cell(ws, "B3", "Rademikibart (anti-IL-4Rα mAb) — Acute Exacerbations of Asthma & COPD | Working Valuation & Diligence File", SUBTITLE, border=False)
ws.merge_cells("B3:G3")
set_cell(ws, "B4", "Prepared as an ongoing analyst working file. All figures in USD unless noted. Blue = input/hardcode. Black = formula. Green = link to another tab.", NOTE, border=False, wrap=True)
ws.merge_cells("B4:G4")

r = 6
set_cell(ws, f"B{r}", "Data as-of date:", BOLD)
set_cell(ws, f"C{r}", "1-Jul-2026 (price/mkt data); 31-Mar-2026 (balance sheet, 10-Q)", BLACK, border=False)
r += 1
set_cell(ws, f"B{r}", "Ticker / Exchange:", BOLD)
set_cell(ws, f"C{r}", "CNTB / Nasdaq Global Market", BLACK)
r += 1
set_cell(ws, f"B{r}", "Lead asset:", BOLD)
set_cell(ws, f"C{r}", "Rademikibart (anti-IL-4Rα mAb)", BLACK)
ws.merge_cells(f"C{r}:F{r}")

r += 2
PRICE_ROW = r
set_cell(ws, f"B{r}", "Current Share Price ($)", BOLD)
set_cell(ws, f"C{r}", 2.40, BLUE, FILL_INPUT, CUR2)
set_cell(ws, f"G{r}", "Source: StockAnalysis.com / Morningstar, intraday ~$2.35-2.45 late Jun/1-Jul 2026. Stock is highly volatile (52-wk range $1.23-$3.82) — UPDATE before relying on this file.", NOTE, border=False, wrap=True)

r += 1
SHARES_ROW = r
set_cell(ws, f"B{r}", "Shares Outstanding, basic (mm)", BOLD)
set_cell(ws, f"C{r}", 62.96, BLUE, FILL_INPUT, NUM1)
set_cell(ws, f"G{r}", "Source: Morningstar/StockAnalysis.com, ~Jul 2026, reflects Mar-2026 6.13mm-share private placement. Q1'26 10-Q weighted-avg (basic/diluted) was 56.5mm.", NOTE, border=False, wrap=True)

r += 1
FD_ROW = r
set_cell(ws, f"B{r}", "Est. Fully-Diluted Shares (mm)", BOLD)
set_cell(ws, f"C{r}", 66.5, BLUE, FILL_INPUT, NUM1)
set_cell(ws, f"G{r}", "ESTIMATE — GENUINE UNKNOWN. We do not have the FY2025 10-K/proxy option, RSU, and warrant tables. Placeholder assumes ~+5.6% dilution from equity awards, in line with typical micro-cap biotech burn-rate. Resolve by pulling the FY2025 Form 10-K equity compensation table or the most recent proxy.", RED, border=False, wrap=True)

r += 1
MKTCAP_ROW = r
set_cell(ws, f"B{r}", "Market Cap, basic ($mm)", BOLD)
set_cell(ws, f"C{r}", f"=C{PRICE_ROW}*C{SHARES_ROW}", BLACK, num_fmt=CUR)

r += 1
set_cell(ws, f"B{r}", "52-Week Range ($)", BOLD)
set_cell(ws, f"C{r}", "1.23 - 3.82", BLACK)

r += 1
CASH_ROW = r
set_cell(ws, f"B{r}", "Cash, Equiv. & ST Investments ($mm)", BOLD)
set_cell(ws, f"C{r}", "='Balance Sheet & Cash Runway'!C10", GREEN, num_fmt=CUR)

r += 1
DEBT_ROW = r
set_cell(ws, f"B{r}", "Total Debt ($mm)", BOLD)
set_cell(ws, f"C{r}", "='Balance Sheet & Cash Runway'!C11", GREEN, num_fmt=CUR)

r += 1
EV_ROW = r
set_cell(ws, f"B{r}", "Enterprise Value, basic shares ($mm)", BOLD)
set_cell(ws, f"C{r}", f"=C{MKTCAP_ROW}-C{CASH_ROW}+C{DEBT_ROW}", BLACK, num_fmt=CUR)
set_cell(ws, f"G{r}", "EV is negative at the current price/cash level — a common feature of clinical-stage biotech trading near or below net cash, and worth flagging as either deep value or a market signal on program risk.", NOTE, border=False, wrap=True)

wb.save("/home/claude/cntb/model.xlsx")
print("Part 1 saved. Next row:", r+2)
print("PRICE_ROW", PRICE_ROW, "SHARES_ROW", SHARES_ROW, "FD_ROW", FD_ROW, "MKTCAP_ROW", MKTCAP_ROW, "CASH_ROW", CASH_ROW, "DEBT_ROW", DEBT_ROW, "EV_ROW", EV_ROW)
