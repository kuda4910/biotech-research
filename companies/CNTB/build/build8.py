import openpyxl, json
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                      HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                      FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

row_map = json.load(open("/home/claude/cntb/row_map.json"))

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb.create_sheet("Price Target Summary")
ws.sheet_view.showGridLines = False

widths = {"A": 3, "B": 34, "C": 15, "D": 13, "E": 15, "F": 50, "G": 3}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

set_cell(ws, "B2", "Price Target Summary", TITLE, border=False)
ws.merge_cells("B2:F2")
set_cell(ws, "B3", "Blends the risk-adjusted (rNPV) and relative-valuation methods with editable weights. The DCF ceiling case is shown for reference, not blended in — including an unrisked 'if it succeeds' number in a probability-weighted target would double-count upside.", SUBTITLE, border=False, wrap=True)
ws.merge_cells("B3:F3")
ws.row_dimensions[3].height = 28

r = 5
set_cell(ws, f"B{r}", "Current share price ($)", BOLD_BLACK)
set_cell(ws, f"C{r}", "='Cover & Assumptions'!C10", GREEN, num_fmt=CUR2)
CUR_PRICE_ROW = r; r += 2

section_bar(ws, r, "METHOD BLEND (Base case scenario)"); r += 2
hdr = r
for i, h in enumerate(["Method", "Value / share ($)", "Weight", "Weighted contribution"]):
    col = chr(ord('B')+i)
    set_cell(ws, f"{col}{hdr}", h, HEADER, FILL_HEADER, align="center")
r += 1

set_cell(ws, f"B{r}", "rNPV — bottom-up, risk-adjusted (Base)", BLACK)
set_cell(ws, f"C{r}", f"='rNPV Model'!C{row_map['RNPV_PER_SHARE_ROW']}", GREEN, num_fmt=CUR2)
set_cell(ws, f"D{r}", 0.60, BLUE, FILL_INPUT, PCT, align="center")
RNPV_ROW = r; r += 1

set_cell(ws, f"B{r}", "Relative valuation — EV / risk-adj. peak sales", BLACK)
set_cell(ws, f"C{r}", f"='Relative Valuation'!C{row_map['REL_PER_SHARE_ROW']}", GREEN, num_fmt=CUR2)
set_cell(ws, f"D{r}", 0.40, BLUE, FILL_INPUT, PCT, align="center")
REL_ROW = r; r += 1

for rr in [RNPV_ROW, REL_ROW]:
    set_cell(ws, f"E{rr}", f"=C{rr}*D{rr}", BLACK, num_fmt=CUR2)

set_cell(ws, f"B{r}", "Weight check (should = 100%)", NOTE, border=False)
set_cell(ws, f"D{r}", f"=D{RNPV_ROW}+D{REL_ROW}", BLACK, num_fmt=PCT)
r += 1

set_cell(ws, f"B{r}", "BLENDED PRICE TARGET (Base case, ~12mo)", BOLD_BLACK, FILL_YELLOW)
set_cell(ws, f"E{r}", f"=E{RNPV_ROW}+E{REL_ROW}", BOLD_BLACK, FILL_YELLOW, num_fmt=CUR2)
BLEND_ROW = r; r += 1

set_cell(ws, f"B{r}", "Implied upside / (downside) vs. current price", BOLD_BLACK)
set_cell(ws, f"E{r}", f"=E{BLEND_ROW}/C{CUR_PRICE_ROW}-1", BOLD_BLACK, num_fmt=PCT)
r += 2

set_cell(ws, f"B{r}", "For reference, not blended:", BOLD_BLACK, border=False)
r += 1
set_cell(ws, f"B{r}", "DCF ceiling case (\"if it succeeds\")", BLACK)
set_cell(ws, f"C{r}", f"='DCF Model'!C{row_map['DCF_PER_SHARE_ROW']}", GREEN, num_fmt=CUR2)
r += 1
set_cell(ws, f"B{r}", "rNPV — Low scenario", BLACK)
set_cell(ws, f"C{r}", "='rNPV Model'!C" + str(row_map['RNPV_PER_SHARE_ROW']) + "", GREEN, num_fmt=CUR2)
set_cell(ws, f"F{r}", "NOTE: this cell reads the Selected-scenario rNPV output, which only reflects Low when the Cover tab's scenario selector is set to Low. It will not update independently — flip the selector to see it live.", NOTE, border=False, wrap=True)
r += 1
set_cell(ws, f"B{r}", "Street consensus target (per sell-side, Jul-2026)", BLACK)
set_cell(ws, f"C{r}", 7.00, BLUE, FILL_INPUT, CUR2)
set_cell(ws, f"F{r}", "Median of 7 analysts per StockAnalysis.com (Jul-2026): Wainwright $7 (Buy), BTIG $10 (Buy), Piper Sandler $7 (Overweight), Oppenheimer $8 (Outperform), Canaccord $6 (Buy), Cantor Fitzgerald $4 (Overweight). See Sell-Side Reconciliation section of the memo.", NOTE, border=False, wrap=True)
r += 2

section_bar(ws, r, "SENSITIVITY — WHAT MOVES THE NUMBER MOST (one input flexed at a time, rNPV Base case, $/share impact)"); r += 2
hdr2 = r
for i, h in enumerate(["Driver", "Downside flex", "Upside flex", "Notes"]):
    col = chr(ord('B')+i)
    set_cell(ws, f"{col}{hdr2}", h, HEADER, FILL_HEADER, align="center")
r += 1

sens = [
 ("Net price per treated patient (±25%)", -1.33, 1.33, "Largest single lever — and the one we have the least real data on. First pricing signal will likely come from post-approval payer commentary, years away; near-term, watch for any investor-day pricing hints."),
 ("Peak penetration of eligible population (±5pp)", -1.33, 1.33, "Tied with price for largest impact — reflects both efficacy differentiation and how fast ED/urgent-care physicians adopt a biologic at point of acute care."),
 ("WACC / discount rate (±2pp)", -0.88, 1.11, "Meaningful given the long dated (2030-32+) launch — a re-rating of biotech risk premia broadly would move this without any CNTB-specific news."),
 ("PoS: COPD approval (±10pp)", -0.77, 0.77, "COPD is intentionally the higher-risk, higher-uncertainty program in this file (see Assumptions tab notes) — a COPD-specific miss or hold would show up here first."),
 ("PoS: Asthma Phase 2 positive readout (±10pp)", -0.41, 0.41, "Smaller than it may look ONLY because it's partially offset by the Ph2→approval conversion probability multiplying it down regardless; the September 2026 readout itself is still the single highest-stakes near-term event — a negative readout removes the Asthma program (~$164mm of Base-case rNPV) almost entirely, not just this incremental slice."),
]
for name, dn, up, note in sens:
    set_cell(ws, f"B{r}", name, BLACK, wrap=True)
    set_cell(ws, f"C{r}", dn, BLACK, num_fmt='+$0.00;-$0.00')
    set_cell(ws, f"D{r}", up, BLACK, num_fmt='+$0.00;-$0.00')
    set_cell(ws, f"E{r}", note, NOTE, border=False, wrap=True)
    ws.merge_cells(f"E{r}:F{r}")
    ws.row_dimensions[r].height = 46
    r += 1
r += 1

set_cell(ws, f"B{r}", ("Read this table as: holding everything else at Base, flexing ONE input by the stated amount moves rNPV/share by the amount shown. Because the drivers interact "
                       "(the Low/High columns on the Cover tab move several at once), the full Low-to-High spread on the rNPV tab ($0.91-$46.47) is much wider than the sum of these one-at-a-time moves — "
                       "that gap is itself the honest message: this stock's value is dominated by compounding uncertainty, not any single number."), BLACK, border=False, wrap=True)
ws.merge_cells(f"B{r}:F{r+2}")
ws.row_dimensions[r].height = 55

wb.save("/home/claude/cntb/model.xlsx")
print("done, next row", r)
