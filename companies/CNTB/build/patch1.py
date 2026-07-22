import openpyxl, json
row_map = json.load(open("/home/claude/cntb/row_map.json"))
wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb["Cover & Assumptions"]
ws[f"C{row_map['CASH_ROW']}"] = f"='Balance Sheet & Cash Runway'!C{row_map['BS_TOTAL_CASH_ROW']}"
ws[f"C{row_map['DEBT_ROW']}"] = f"='Balance Sheet & Cash Runway'!C{row_map['BS_DEBT_ROW']}"
wb.save("/home/claude/cntb/model.xlsx")
print("patched")
