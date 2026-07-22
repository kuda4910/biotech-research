import openpyxl, json
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                     HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                     FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb.create_sheet("Balance Sheet & Cash Runway")

widths = {"A": 3, "B": 40, "C": 15, "D": 15, "E": 15, "F": 15, "G": 45, "H": 3}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.sheet_view.showGridLines = False

set_cell(ws, "B2", "Balance Sheet & Cash Runway", TITLE, border=False)
ws.merge_cells("B2:F2")
set_cell(ws, "B3", "Source: Form 10-Q for the quarter ended 31-Mar-2026 (filed 12-May-2026) + 30-Mar-2026 private placement 8-K. Cash items below are the last reported balance sheet; UPDATE each quarter.", SUBTITLE, border=False, wrap=True)
ws.merge_cells("B3:G3")

r = 5
section_bar(ws, r, "CASH POSITION (as of 31-Mar-2026, $mm)"); r += 2
rows = [
 ("Cash and cash equivalents", 30.1, "Implied: total cash & ST invest. of $46.0mm less $15.9mm ST investments (commercial paper) per Q1'26 10-Q. VERIFY exact split against the filed balance sheet."),
 ("Short-term investments (US commercial paper)", 15.9, "Per Q1'26 10-Q liquidity disclosure."),
]
CASH_EQ_ROW = r
set_cell(ws, f"B{r}", rows[0][0], BLACK); set_cell(ws, f"C{r}", rows[0][1], BLUE, FILL_INPUT, CUR); set_cell(ws, f"G{r}", rows[0][2], NOTE, border=False, wrap=True); r+=1
ST_INV_ROW = r
set_cell(ws, f"B{r}", rows[1][0], BLACK); set_cell(ws, f"C{r}", rows[1][1], BLUE, FILL_INPUT, CUR); set_cell(ws, f"G{r}", rows[1][2], NOTE, border=False, wrap=True); r+=1
TOTAL_CASH_ROW = r
set_cell(ws, f"B{r}", "Total cash, cash equivalents & ST investments", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=C{CASH_EQ_ROW}+C{ST_INV_ROW}", BOLD_BLACK, num_fmt=CUR)
set_cell(ws, f"G{r}", "Company-reported total: $46.0mm as of 31-Mar-2026. This should tie out — if the two-line split above is later revised from the actual balance sheet, this total is the control check.", NOTE, border=False, wrap=True)
r += 2

DEBT_ROW = r
section_bar(ws, r-1, "") # no-op spacer already added; continue
set_cell(ws, f"B{r}", "Total debt", BOLD_BLACK)
set_cell(ws, f"C{r}", 0, BLUE, FILL_INPUT, CUR)
set_cell(ws, f"G{r}", "Company reported zero debt as of the last balance sheet (Simply Wall St / 10-Q). No covenants, tranching, or maturity schedule to disclose — capital structure is currently all-equity.", NOTE, border=False, wrap=True)
r += 2

section_bar(ws, r, "RECENT FINANCING & DILUTION"); r += 2
fin_rows = [
 ("30-Mar-2026 private placement — gross proceeds ($mm)", 20.2, "6,130,000 ordinary shares at $3.25/share to new & existing investors. Closed 31-Mar-2026."),
 ("30-Mar-2026 private placement — est. net proceeds ($mm)", 18.6, "Per company release; ~8% placement/legal costs implied."),
 ("Director open-market buy, 29-May-2026 ($mm)", 0.37, "Director James Huang bought 150,000 sh @ $2.48 — open-market, i.e., a real conviction signal (not options/RSU mechanics)."),
 ("Director PIPE participation, 31-Mar-2026 ($mm)", 4.0, "Director James Huang bought 1.16mm sh in the private placement itself (~$3.25-3.45/sh) — this is PIPE participation at a negotiated financing price, not an open-market buy; read as supportive but not the same signal as rows above."),
]
for label, val, note in fin_rows:
    set_cell(ws, f"B{r}", label, BLACK, wrap=True)
    set_cell(ws, f"C{r}", val, BLUE, FILL_INPUT, CUR2)
    set_cell(ws, f"G{r}", note, NOTE, border=False, wrap=True)
    ws.row_dimensions[r].height = 30
    r += 1
r += 1

section_bar(ws, r, "SHARE COUNT — BASIC VS. FULLY DILUTED"); r += 2
sc_rows = [
 ("Ordinary shares outstanding, basic (mm)", 62.96, "Current, post private placement. Source: company disclosures / StockAnalysis.com."),
 ("Q1 2026 weighted-avg shares, basic & diluted (mm)", 56.5, "Per 10-Q EPS calc — lower than current count because the 6.13mm PIPE shares closed on the last day of the quarter."),
 ("Est. options / RSUs outstanding (mm) — UNKNOWN", 3.5, "GENUINE UNKNOWN — we do not have the FY2025 10-K equity-comp table. Placeholder only."),
]
for label, val, note in sc_rows:
    set_cell(ws, f"B{r}", label, BLACK, wrap=True)
    set_cell(ws, f"C{r}", val, BLUE, FILL_INPUT, NUM1)
    set_cell(ws, f"G{r}", note, RED if "UNKNOWN" in note else NOTE, border=False, wrap=True)
    ws.row_dimensions[r].height = 28
    r += 1
r += 1

section_bar(ws, r, "BURN RATE & CASH RUNWAY"); r += 2
burn_rows = [
 ("Q1 2026 R&D expense ($mm)", 15.0),
 ("Q1 2026 G&A expense ($mm)", 4.7),
 ("Q1 2026 license/collaboration revenue ($mm)", 0.169),
 ("Q1 2026 net loss ($mm)", 19.4),
 ("Q1 2026 operating cash outflow ($mm)", 16.0),
]
Q1_OPEX_START = r
for label, val in burn_rows:
    set_cell(ws, f"B{r}", label, BLACK)
    set_cell(ws, f"C{r}", val, BLUE, FILL_INPUT, CUR2)
    r += 1
set_cell(ws, f"G{Q1_OPEX_START}", "Source: Q1 2026 10-Q / 8-K (filed 12-May-2026). R&D more than doubled y/y as SEABREEZE STAT enrollment ramped; net loss widened to $19.4mm ($0.34/sh) from $10.3mm ($0.19/sh) in Q1 2025.", NOTE, border=False, wrap=True)

QUARTERLY_BURN_ROW = r
set_cell(ws, f"B{r}", "Assumed steady-state quarterly cash burn, next 12mo ($mm)", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=C{Q1_OPEX_START+4}*1.05", BLACK, num_fmt=CUR)
set_cell(ws, f"G{r}", "ESTIMATE: Q1'26 actual op. cash outflow +5%, reflecting COPD enrollment completion and ramp toward two Ph2 data readouts. Update once Q2 2026 10-Q (due ~12-Aug-2026) is out.", NOTE, border=False, wrap=True)
r += 2

RUNWAY_Q_ROW = r
set_cell(ws, f"B{r}", "Implied quarters of runway from current cash", BOLD_BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C15/C{QUARTERLY_BURN_ROW}", BLACK, num_fmt='0.0" qtrs"')
r += 1
RUNWAY_DATE_ROW = r
set_cell(ws, f"B{r}", "Implied runway exhaustion (quarter-end basis)", BOLD_BLACK)
set_cell(ws, f"C{r}", f'=EDATE(DATE(2026,3,31),3*C{RUNWAY_Q_ROW})', BLACK, num_fmt='mmm-yyyy')
set_cell(ws, f"G{r}", "Management's own guidance is cash runway 'into 2H 2027' (i.e., through at least Sep-2027) — a bit longer than the mechanical 4-quarter-burn extrapolation above, likely because management is netting in expected cost discipline / does not assume acceleration into Ph3. Treat the two figures as a range, not a single point estimate.", NOTE, border=False, wrap=True)
r += 2

section_bar(ws, r, "FINANCING NEED / DILUTION RISK FLAG"); r += 2
set_cell(ws, f"B{r}", ("SEABREEZE STAT data (Sept 2026) precedes the low end of the mechanical runway estimate. If both readouts are positive, CNTB will almost certainly need to raise capital "
                       "to fund a Phase 3 program — likely via a follow-on equity raise, a royalty/synthetic-royalty deal (as Apogee Therapeutics did with AbbVie/undisclosed lender), or a Simcere-style ex-China licensing deal. "
                       "Any of these is dilutive or cash-flow-negative in the near term and should be modeled explicitly once terms are known — none are in this file yet."), BLACK, border=False, wrap=True)
ws.merge_cells(f"B{r}:G{r+2}")
ws.row_dimensions[r].height = 60

ws.freeze_panes = "B5"
wb.save("/home/claude/cntb/model.xlsx")

row_map = json.load(open("/home/claude/cntb/row_map.json"))
row_map.update({
    "BS_TOTAL_CASH_ROW": TOTAL_CASH_ROW, "BS_DEBT_ROW": DEBT_ROW, "BS_RUNWAY_Q_ROW": RUNWAY_Q_ROW,
})
json.dump(row_map, open("/home/claude/cntb/row_map.json","w"))
print(row_map)
