"""
Patch 7: revise the Loss-of-exclusivity assumption (Cover & Assumptions row 52) with actual
patent data from CNTB's FY2024 and FY2025 Form 10-Ks (Item 1, Intellectual Property), replacing
the placeholder guess flagged in the original source note as unverified.

Source facts (stable across both the FY2024 10-K, filed 2025-03-31, and the FY2025 10-K, filed
2026-03-31):
- Composition-of-matter patent family for rademikibart: expected to expire 2037, not accounting
  for patent term adjustment, extensions, or terminal disclaimers.
- Formulation patent family: expected to expire 2040, same caveats.
- Hatch-Waxman Act patent term extension (US): up to 5 years, capped at 14 years total from the
  date of product approval; only one patent per approved product may be extended.

Revised Low/Base/High:
- Low = 2037 (composition-of-matter family, unextended -- the earliest-controlling patent, worst
  case for biosimilar entry timing)
- Base = 2040 (formulation family, unextended -- unchanged from the prior placeholder, now sourced
  instead of guessed)
- High = 2045 (formulation family + a plausible Hatch-Waxman extension, consistent with the
  ~2031 base-case launch year elsewhere in this tab plus the 14-years-from-approval cap)

IMPORTANT: openpyxl does not recalculate formulas -- it only updates the input cells. The
workbook's cached rNPV/DCF/blended outputs will be stale until this file is opened in Excel (or
recalculated with LibreOffice headless) at least once. Do not read Low/Base/High rNPV figures out
of this file's cached values until that's been done; recompute portfolio.json's bear/mid/bull
only after a real recalc, not from this patch.
"""
import openpyxl

PATH = r"C:\Users\penel\Desktop\biotech-research\companies\CNTB\CNTB_Valuation_Model.xlsx"

wb = openpyxl.load_workbook(PATH, data_only=False)
ws = wb["Cover & Assumptions"]

ws["C52"] = 2037
ws["D52"] = 2040
ws["E52"] = 2045
ws["G52"] = (
    "SOURCED (was an unverified placeholder). Per FY2024 10-K (filed 2025-03-31) and FY2025 "
    "10-K (filed 2026-03-31), Item 1 Intellectual Property: composition-of-matter patent family "
    "expected to expire 2037 (unextended); formulation family expected to expire 2040 "
    "(unextended); Hatch-Waxman PTE could add up to 5yr, capped at 14yr from approval. "
    "Low=2037 (composition-of-matter, unextended). Base=2040 (formulation, unextended, unchanged "
    "from prior placeholder). High=2045 (formulation + plausible PTE, consistent with the ~2031 "
    "base-case launch year and the 14-year-from-approval cap)."
)

wb.save(PATH)
print("patch7 applied: row 52 (Loss of exclusivity year) updated to Low=2037 Base=2040 High=2045")
