import openpyxl, json
from openpyxl.utils import get_column_letter
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                      HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                      FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

row_map = json.load(open("/home/claude/cntb/row_map.json"))
N = row_map["N_YEARS"]; FIRST_COL_IDX = row_map["FIRST_COL_IDX"]
def yc(i): return get_column_letter(FIRST_COL_IDX + i)
RM = "'rNPV Model'"
REV_A_ROW = row_map["REV_A_ROW"]; REV_C_ROW = row_map["REV_C_ROW"]
COMBPOS_A_ROW = row_map["COMBPOS_A_ROW"]; COMBPOS_C_ROW = row_map["COMBPOS_C_ROW"]

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb.create_sheet("Relative Valuation")
ws.sheet_view.showGridLines = False

widths = {"A": 3, "B": 34, "C": 16, "D": 16, "E": 16, "F": 16, "G": 50, "H": 3}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

set_cell(ws, "B2", "Relative Valuation — EV / Risk-Adjusted Peak Sales", TITLE, border=False)
ws.merge_cells("B2:G2")
set_cell(ws, "B3", ("Genuine peer set for a single-asset, Phase 2, respiratory/T2-inflammation biotech is thin — most direct comps have either been acquired or are far more advanced. "
                     "We use one active public peer (Upstream Bio) for a market-derived multiple, and two precedent transactions as sanity checks, not as the primary multiple source."), SUBTITLE, border=False, wrap=True)
ws.merge_cells("B3:G3")
ws.row_dimensions[3].height = 28

r = 5
section_bar(ws, r, "PEER SCREEN & INCLUSION/EXCLUSION LOGIC"); r += 2

peers = [
 ("Upstream Bio (UPB)", "INCLUDED — primary multiple source",
  "Single-asset (verekitug, TSLP-receptor antagonist), Phase 2/3, severe asthma + CRSwNP + COPD. Directly comparable stage/business model: pre-revenue, single-mechanism, T2-respiratory biologic, similarly exposed to one binary readout at a time (UPB fell ~47% on its own Phase 2 VALIANT asthma miss in Feb 2026 — a useful real-world illustration of CNTB's own Sept-2026 risk)."),
 ("Aiolos Bio / GSK", "REFERENCE ONLY — precedent transaction, not a multiple",
  "GSK acquired Aiolos (Phase 2-ready TSLP mAb, asthma-only, Jan-2024) for $1.0bn upfront + up to $400mm in milestones. This is a control/M&A price, not a market-trading multiple, and covers a single indication (asthma) vs. CNTB's two (asthma+COPD) — used below only as a directional sanity check, not blended into the multiple."),
 ("Apogee Therapeutics / AbbVie", "EXCLUDED from multiple",
  "AbbVie's ~$10.9bn acquisition (2026) is real and topical, but Apogee is a multi-asset platform (zumilokibart in AD/asthma/EoE + APG222 + more) with an $800mm synthetic-royalty and $500mm debt facility already in place and cash into 2029 — a materially larger, more advanced, more de-risked company. Applying its scale of multiple to a single Phase-2 asset like CNTB would be a scale artifact, not a valid read-across. Mentioned only as evidence that large pharma has active, recent appetite for the IL-4/IL-13/T2 respiratory space."),
]
for name, status, note in peers:
    set_cell(ws, f"B{r}", name, BOLD_BLACK)
    set_cell(ws, f"C{r}", status, BLACK, wrap=True)
    ws.merge_cells(f"C{r}:D{r}")
    set_cell(ws, f"E{r}", note, NOTE, border=False, wrap=True)
    ws.merge_cells(f"E{r}:G{r}")
    ws.row_dimensions[r].height = 55
    r += 1
r += 1

section_bar(ws, r, "PEER 1 — UPSTREAM BIO (UPB): MULTIPLE DERIVATION"); r += 2
rows1 = [
 ("UPB enterprise value ($mm)", 190, "Seeking Alpha initiation (30-May-2026) cites UPB trading at 'an enterprise value under $200M' following the post-VALIANT-data selloff; we use $190mm as a point estimate. VERIFY against a current 10-Q cash balance and market cap before relying on this."),
 ("UPB risk-adjusted peak WW sales estimate ($mm, 2035)", 2700, "Mizuho equity research (17-Dec-2025): '$2.7 billion in risk-adjusted worldwide sales by 2035' for verekitug across asthma + CRSwNP + COPD combined. This is already probability-weighted, not an unrisked peak-sales figure — important for how we apply the resulting multiple below."),
]
UPB_EV_ROW = r
set_cell(ws, f"B{r}", rows1[0][0], BLACK, wrap=True); set_cell(ws, f"C{r}", rows1[0][1], BLUE, FILL_INPUT, CUR); set_cell(ws, f"E{r}", rows1[0][2], NOTE, border=False, wrap=True); ws.merge_cells(f"E{r}:G{r}"); ws.row_dimensions[r].height=42; r+=1
UPB_PKSALES_ROW = r
set_cell(ws, f"B{r}", rows1[1][0], BLACK, wrap=True); set_cell(ws, f"C{r}", rows1[1][1], BLUE, FILL_INPUT, CUR); set_cell(ws, f"E{r}", rows1[1][2], NOTE, border=False, wrap=True); ws.merge_cells(f"E{r}:G{r}"); ws.row_dimensions[r].height=42; r+=1

UPB_MULT_ROW = r
set_cell(ws, f"B{r}", "Implied EV / risk-adjusted peak sales multiple", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=C{UPB_EV_ROW}/C{UPB_PKSALES_ROW}", BOLD_BLACK, num_fmt=MULT)
set_cell(ws, f"E{r}", "Note the basis: this multiple is EV per dollar of RISK-ADJUSTED (already probability-weighted) peak sales — not unrisked peak sales. We must apply it to CNTB's risk-adjusted peak sales figure below, not the unrisked one, or we would be double- or under-counting clinical risk (a 'share-count artifact' -style error in the other direction).", NOTE, border=False, wrap=True)
ws.merge_cells(f"E{r}:G{r}"); ws.row_dimensions[r].height = 55
r += 2

section_bar(ws, r, "CNTB — RISK-ADJUSTED PEAK SALES (SAME BASIS AS THE PEER MULTIPLE)"); r += 2

set_cell(ws, f"B{r}", "Peak unrisked US net revenue — Asthma ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=MAX({RM}!{yc(0)}{REV_A_ROW}:{yc(N-1)}{REV_A_ROW})", GREEN, num_fmt=CUR)
CNTB_PK_A_ROW = r; r += 1

set_cell(ws, f"B{r}", "Peak unrisked US net revenue — COPD ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=MAX({RM}!{yc(0)}{REV_C_ROW}:{yc(N-1)}{REV_C_ROW})", GREEN, num_fmt=CUR)
CNTB_PK_C_ROW = r; r += 1

set_cell(ws, f"B{r}", "Combined PoS to approval — Asthma", BLACK)
set_cell(ws, f"C{r}", f"={RM}!$C${COMBPOS_A_ROW}", GREEN, num_fmt=PCT)
CNTB_POS_A_ROW = r; r += 1

set_cell(ws, f"B{r}", "Combined PoS to approval — COPD", BLACK)
set_cell(ws, f"C{r}", f"={RM}!$C${COMBPOS_C_ROW}", GREEN, num_fmt=PCT)
CNTB_POS_C_ROW = r; r += 1

set_cell(ws, f"B{r}", "CNTB risk-adjusted peak sales ($mm)", BOLD_BLACK, FILL_LIGHT)
set_cell(ws, f"C{r}", f"=C{CNTB_PK_A_ROW}*C{CNTB_POS_A_ROW}+C{CNTB_PK_C_ROW}*C{CNTB_POS_C_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
CNTB_RAPKS_ROW = r; r += 2

section_bar(ws, r, "IMPLIED VALUE — EV / PEAK-SALES METHOD"); r += 2
set_cell(ws, f"B{r}", "Implied CNTB enterprise value ($mm)", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=C{UPB_MULT_ROW}*C{CNTB_RAPKS_ROW}", BOLD_BLACK, num_fmt=CUR)
REL_EV_ROW = r; r += 1
set_cell(ws, f"B{r}", "Plus: net cash ($mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['CASH_ROW']}-'Cover & Assumptions'!C{row_map['DEBT_ROW']}", GREEN, num_fmt=CUR)
REL_NETCASH_ROW = r; r += 1
set_cell(ws, f"B{r}", "Implied CNTB equity value ($mm)", BOLD_BLACK, FILL_LIGHT)
set_cell(ws, f"C{r}", f"=C{REL_EV_ROW}+C{REL_NETCASH_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
REL_EQ_ROW = r; r += 1
set_cell(ws, f"B{r}", "Fully-diluted shares (mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['FD_ROW']}", GREEN, num_fmt=NUM1)
REL_FD_ROW = r; r += 2

set_cell(ws, f"B{r}", "Relative valuation — value per share ($)", BOLD_BLACK, FILL_YELLOW)
set_cell(ws, f"C{r}", f"=C{REL_EQ_ROW}/C{REL_FD_ROW}", BOLD_BLACK, FILL_YELLOW, num_fmt=CUR2)
REL_PER_SHARE_ROW = r; r += 2

section_bar(ws, r, "PRECEDENT TRANSACTION SANITY CHECKS (not used in the multiple above)"); r += 2
prec = [
 ("Aiolos Bio / GSK, Jan-2024", "$1.0bn upfront + up to $400mm milestones ($1.4bn total)", "Phase 2-ready, single TSLP mAb, asthma only (+ CRSwNP option). If CNTB's dual-indication (asthma+COPD), Phase-2-de-risked-by-Sept-2026 rademikibart commanded even a similar per-indication control price, that alone implies a value well above CNTB's current ~$150mm market cap — a rough upper-bound sanity check on the downside case, not a target."),
 ("Apogee Therapeutics / AbbVie, 2026", "$10.9bn all-cash", "Confirms continued large-pharma appetite for the IL-4/IL-13/T2 space, but Apogee is a multi-program, more clinically advanced platform — not scalable down to a single-asset Phase 2 comp."),
]
for name, terms, note in prec:
    set_cell(ws, f"B{r}", name, BOLD_BLACK)
    set_cell(ws, f"C{r}", terms, BLACK, wrap=True)
    ws.merge_cells(f"C{r}:D{r}")
    set_cell(ws, f"E{r}", note, NOTE, border=False, wrap=True)
    ws.merge_cells(f"E{r}:G{r}")
    ws.row_dimensions[r].height = 60
    r += 1

r += 1
set_cell(ws, f"B{r}", ("Caution on this whole tab: with one true public peer, the 'multiple' above is a single data point, not a statistically robust market multiple — treat it as directionally "
                       "informative, weight it accordingly in the blend on the Price Target Summary tab, and revisit the moment a second respiratory Phase 2 peer trades publicly or a transaction closes at a disclosed price."), BLACK, border=False, wrap=True)
ws.merge_cells(f"B{r}:G{r+2}")
ws.row_dimensions[r].height = 55

wb.save("/home/claude/cntb/model.xlsx")

row_map.update(dict(
    REL_PER_SHARE_ROW=REL_PER_SHARE_ROW, CNTB_RAPKS_ROW=CNTB_RAPKS_ROW, UPB_MULT_ROW=UPB_MULT_ROW,
))
json.dump(row_map, open("/home/claude/cntb/row_map.json","w"))
print("Relative valuation per-share row:", REL_PER_SHARE_ROW)
