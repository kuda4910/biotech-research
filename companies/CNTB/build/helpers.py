import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

FONT = "Arial"

BLUE = Font(name=FONT, color="0000FF", size=10)
BLACK = Font(name=FONT, color="000000", size=10)
GREEN = Font(name=FONT, color="008000", size=10)
BOLD = Font(name=FONT, bold=True, size=10)
BOLD_BLACK = Font(name=FONT, bold=True, color="000000", size=10)
TITLE = Font(name=FONT, bold=True, size=16, color="1F3864")
SUBTITLE = Font(name=FONT, italic=True, size=10, color="595959")
HEADER = Font(name=FONT, bold=True, size=11, color="FFFFFF")
SECTION = Font(name=FONT, bold=True, size=12, color="FFFFFF")
NOTE = Font(name=FONT, italic=True, size=9, color="808080")
RED = Font(name=FONT, color="FF0000", size=10)

FILL_HEADER = PatternFill("solid", fgColor="1F3864")
FILL_SECTION = PatternFill("solid", fgColor="2E5395")
FILL_YELLOW = PatternFill("solid", fgColor="FFFF00")
FILL_LIGHT = PatternFill("solid", fgColor="D9E1F2")
FILL_GREY = PatternFill("solid", fgColor="F2F2F2")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
CUR2 = '$#,##0.00;($#,##0.00);"-"'
PCT = '0.0%'
MULT = '0.0x'
NUM = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'

def set_cell(ws, coord, value=None, font=None, fill=None, num_fmt=None, align=None, border=True, wrap=False, comment=None):
    c = ws[coord]
    if value is not None:
        c.value = value
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if num_fmt:
        c.number_format = num_fmt
    if align:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    elif wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    if border:
        c.border = BORDER
    if comment:
        c.comment = Comment(comment, "Analyst")
    return c

def section_bar(ws, row, text, col_start=1, col_end=8):
    for col in range(col_start, col_end+1):
        cell = ws.cell(row=row, column=col)
        cell.fill = FILL_SECTION
        cell.font = SECTION
    ws.cell(row=row, column=col_start).value = text

def sel_formula(cell_low, cell_high_range_ref, scenario_cell="$B$29"):
    pass

