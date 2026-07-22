import openpyxl, json
from openpyxl.utils import get_column_letter
from helpers import (set_cell, section_bar, BLUE, BLACK, GREEN, BOLD, BOLD_BLACK, TITLE, SUBTITLE,
                      HEADER, SECTION, NOTE, RED, FILL_HEADER, FILL_SECTION, FILL_YELLOW, FILL_LIGHT,
                      FILL_GREY, FILL_INPUT, BORDER, CUR, CUR2, PCT, MULT, NUM, NUM1)

row_map = json.load(open("/home/claude/cntb/row_map.json"))
A = row_map["ASSUMP_START"]
WACC = A+13
CA = "'Cover & Assumptions'"
def af(row): return f"{CA}!$F${row}"

N = row_map["N_YEARS"]; FIRST_COL_IDX = row_map["FIRST_COL_IDX"]
def yc(i): return get_column_letter(FIRST_COL_IDX + i)

RM = "'rNPV Model'"
AT_A_ROW = row_map["AT_A_ROW"]
AT_C_ROW = row_map["AT_C_ROW"]
YEAR_ROW = row_map["YEAR_ROW"]

wb = openpyxl.load_workbook("/home/claude/cntb/model.xlsx")
ws = wb.create_sheet("DCF Model")
ws.sheet_view.showGridLines = False

widths = {"A": 3, "B": 46}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
for i in range(N):
    ws.column_dimensions[yc(i)].width = 11
ws.column_dimensions["X"].width = 55

set_cell(ws, "B2", "DCF Model — Unrisked \"If It Succeeds\" Ceiling Case", TITLE, border=False)
set_cell(ws, "B3", ("THIS IS THE CEILING CASE, NOT THE BASE CASE. It assumes both Phase 2 programs succeed and convert to approval with certainty (100% PoS) — it strips out "
                     "the binary clinical/regulatory risk-adjustment that the rNPV tab applies, but keeps the same underlying operating engine (same patient counts, pricing, cost structure, launch timing selected on the Cover tab). "
                     "Use it to see what CNTB is worth if the science works, not as a probability-weighted target."), SUBTITLE, border=False, wrap=True)
ws.merge_cells(f"B3:{yc(N-1)}5")
ws.row_dimensions[3].height = 60

r = 7
set_cell(ws, f"B{r}", "Fiscal Year", HEADER, FILL_HEADER, align="center")
for i in range(N):
    set_cell(ws, f"{yc(i)}{r}", f"={RM}!{yc(i)}{YEAR_ROW}", HEADER, FILL_HEADER, align="center", num_fmt="0")
DCF_YEAR_ROW = r
r += 2

section_bar(ws, r, "UNRISKED OPERATING CASH FLOW (linked from rNPV Model tab — same engine, no PoS haircut)"); r += 2

set_cell(ws, f"B{r}", "Unrisked after-tax cash flow — Asthma ($mm)", BLACK)
DCF_AT_A_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={RM}!{col}{AT_A_ROW}", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Unrisked after-tax cash flow — COPD ($mm)", BLACK)
DCF_AT_C_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={RM}!{col}{AT_C_ROW}", BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Total unrisked after-tax cash flow ($mm)", BOLD_BLACK)
DCF_TOTAL_CF_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{DCF_AT_A_ROW}+{col}{DCF_AT_C_ROW}", BOLD_BLACK, num_fmt=CUR, align="center")
r += 1

set_cell(ws, f"B{r}", "Discount factor (mid-year convention, @ selected WACC)", BLACK)
DCF_DISC_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"=1/(1+{af(WACC)})^({col}{DCF_YEAR_ROW}-2026+0.5)", BLACK, num_fmt='0.000', align="center")
r += 1

set_cell(ws, f"B{r}", "PV of unrisked cash flow ($mm)", BLACK)
DCF_PV_ROW = r
for i in range(N):
    col = yc(i)
    set_cell(ws, f"{col}{r}", f"={col}{DCF_TOTAL_CF_ROW}*{col}{DCF_DISC_ROW}", BLACK, num_fmt=CUR, align="center")
r += 2

set_cell(ws, f"B{r}", "Sum of PV, explicit period 2026-2045 ($mm)", BOLD_BLACK)
DCF_SUM_PV_ROW = r
set_cell(ws, f"C{r}", f"=SUM({yc(0)}{DCF_PV_ROW}:{yc(N-1)}{DCF_PV_ROW})", BOLD_BLACK, num_fmt=CUR)
r += 2

section_bar(ws, r, "TERMINAL VALUE"); r += 2
set_cell(ws, f"B{r}", "Terminal-year (2045) post-LOE steady-state cash flow ($mm)", BLACK)
DCF_TERM_CF_ROW = r
set_cell(ws, f"C{r}", f"={yc(N-1)}{DCF_TOTAL_CF_ROW}", BLACK, num_fmt=CUR)
set_cell(ws, "X"+str(r), "By 2045 both programs are well past the modeled loss-of-exclusivity year in every scenario, so this is already a genericized/eroded residual (~15% of peak), not peak-year cash flow — using peak-year cash flow here would materially overstate the terminal value.", NOTE, border=False, wrap=True)
r += 1

set_cell(ws, f"B{r}", "Terminal growth rate", BLACK)
DCF_TERM_G_ROW = r
set_cell(ws, f"C{r}", 0.0, BLUE, FILL_INPUT, PCT)
set_cell(ws, "X"+str(r), "Set to 0% — a mature, off-patent/biosimilar-eroded revenue stream should not be assumed to grow in perpetuity. This is a conservative but standard convention.", NOTE, border=False, wrap=True)
r += 1

set_cell(ws, f"B{r}", "Terminal value, undiscounted ($mm)", BLACK)
DCF_TERM_VAL_ROW = r
set_cell(ws, f"C{r}", f"=C{DCF_TERM_CF_ROW}*(1+C{DCF_TERM_G_ROW})/({af(WACC)}-C{DCF_TERM_G_ROW})", BLACK, num_fmt=CUR)
r += 1

set_cell(ws, f"B{r}", "PV of terminal value ($mm)", BOLD_BLACK)
DCF_PV_TERM_ROW = r
set_cell(ws, f"C{r}", f"=C{DCF_TERM_VAL_ROW}*{yc(N-1)}{DCF_DISC_ROW}", BOLD_BLACK, num_fmt=CUR)
r += 2

section_bar(ws, r, "DCF SUMMARY & PER-SHARE VALUE (CEILING CASE)"); r += 2
set_cell(ws, f"B{r}", "Sum of PV, explicit period ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=C{DCF_SUM_PV_ROW}", BLACK, num_fmt=CUR)
S1 = r; r += 1
set_cell(ws, f"B{r}", "PV of terminal value ($mm)", BLACK)
set_cell(ws, f"C{r}", f"=C{DCF_PV_TERM_ROW}", BLACK, num_fmt=CUR)
S2 = r; r += 1
set_cell(ws, f"B{r}", "China milestones, unrisked, discounted @ 50% PV credit ($mm)", BLACK)
set_cell(ws, f"C{r}", "=110*0.5", BLACK, num_fmt=CUR)
set_cell(ws, "X"+str(r), "Ceiling-case treatment: assumes half of the $110mm remaining Simcere milestones land, discounted for timing but not for Simcere-execution risk (which the rNPV tab already prices in separately).", NOTE, border=False, wrap=True)
S3 = r; r += 1
set_cell(ws, f"B{r}", "China royalty stream, unrisked NPV ($mm)", BLACK)
set_cell(ws, f"C{r}", f"={af(row_map['ASSUMP_START']+15)}*2", BLACK, num_fmt=CUR)
set_cell(ws, "X"+str(r), "Ceiling case: 2x the risk-adjusted China royalty NPV used in the rNPV tab, as a simple proxy for removing most (not all) of the discount applied there.", NOTE, border=False, wrap=True)
S4 = r; r += 1
set_cell(ws, f"B{r}", "Enterprise value, DCF ceiling case ($mm)", BOLD_BLACK)
set_cell(ws, f"C{r}", f"=SUM(C{S1}:C{S4})", BOLD_BLACK, num_fmt=CUR)
DCF_EV_ROW = r; r += 1
set_cell(ws, f"B{r}", "Plus: net cash ($mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['CASH_ROW']}-'Cover & Assumptions'!C{row_map['DEBT_ROW']}", GREEN, num_fmt=CUR)
DCF_NETCASH_ROW = r; r += 1
set_cell(ws, f"B{r}", "Total equity value, DCF ceiling case ($mm)", BOLD_BLACK, FILL_LIGHT)
set_cell(ws, f"C{r}", f"=C{DCF_EV_ROW}+C{DCF_NETCASH_ROW}", BOLD_BLACK, FILL_LIGHT, num_fmt=CUR)
DCF_TOTAL_EQ_ROW = r; r += 1
set_cell(ws, f"B{r}", "Fully-diluted shares (mm)", BLACK)
set_cell(ws, f"C{r}", f"='Cover & Assumptions'!C{row_map['FD_ROW']}", GREEN, num_fmt=NUM1)
DCF_FD_ROW = r; r += 2

set_cell(ws, f"B{r}", "DCF value per share — CEILING CASE ($)", BOLD_BLACK, FILL_YELLOW)
set_cell(ws, f"C{r}", f"=C{DCF_TOTAL_EQ_ROW}/C{DCF_FD_ROW}", BOLD_BLACK, FILL_YELLOW, num_fmt=CUR2)
DCF_PER_SHARE_ROW = r; r += 2

set_cell(ws, f"B{r}", ("This ceiling case is the upside scenario a bull would cite if SEABREEZE data are clean and rademikibart becomes the first biologic approved for acute asthma/COPD exacerbations. "
                       "It is not risk-adjusted for the still-open Phase 2 binary event in September/Q4 2026, and should never be quoted as a price target on its own — pair it with the rNPV tab's Low/Base/High."), BLACK, border=False, wrap=True)
ws.merge_cells(f"B{r}:{yc(8)}{r+2}")
ws.row_dimensions[r].height = 55

ws.freeze_panes = "C8"
wb.save("/home/claude/cntb/model.xlsx")

row_map.update(dict(
    DCF_YEAR_ROW=DCF_YEAR_ROW, DCF_TOTAL_CF_ROW=DCF_TOTAL_CF_ROW, DCF_PV_ROW=DCF_PV_ROW,
    DCF_SUM_PV_ROW=DCF_SUM_PV_ROW, DCF_TERM_VAL_ROW=DCF_TERM_VAL_ROW, DCF_PV_TERM_ROW=DCF_PV_TERM_ROW,
    DCF_EV_ROW=DCF_EV_ROW, DCF_TOTAL_EQ_ROW=DCF_TOTAL_EQ_ROW, DCF_PER_SHARE_ROW=DCF_PER_SHARE_ROW,
))
json.dump(row_map, open("/home/claude/cntb/row_map.json","w"))
print("DCF per-share row:", DCF_PER_SHARE_ROW)
